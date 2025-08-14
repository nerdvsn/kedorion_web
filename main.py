# main.py — FastAPI backend (LOCAL-ONLY)
# Speichert Bewerbungen lokal: Datei in storage/uploads + Zeile in storage/applications.xlsx
# Kein E‑Mail‑Versand. Robust gegen beschädigte/leer angelegte Excel-Dateien.

import os
from typing import List
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile

from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse, FileResponse

# Optional: reCAPTCHA v3 Prüfung (kann deaktiviert bleiben)
import httpx

# Excel-Logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# ---------------------------
# Konfiguration (Env-Variablen)
# ---------------------------
BASE_DIR = Path(__file__).resolve().parent

# Lokale Speicherung
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", BASE_DIR / "storage" / "uploads"))
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", BASE_DIR / "storage" / "applications.xlsx"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

# Limits & Validierung
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "15"))
ALLOWED_EXTENSIONS = {".pdf", ".doc", ".docx"}

# reCAPTCHA v3 (leer lassen, um zu deaktivieren)
RECAPTCHA_SECRET = os.getenv("RECAPTCHA_SECRET", "")

# ---------------------------
# Utilities
# ---------------------------

def _new_wb_with_header(keys: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(list(keys))
    return wb, ws


def append_to_excel(row: dict):
    """Schreibt oder erzeugt applications.xlsx und hängt eine Bewerbung an.
    Erkennt beschädigte Dateien und legt automatisch neu an (mit Backup)."""
    keys = list(row.keys())

    wb = None
    ws = None

    if EXCEL_PATH.exists():
        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            # Falls Datei leer oder ohne Header ist → Header setzen
            if ws.max_row == 1 and (ws["A1"].value is None):
                ws.append(keys)
        except (InvalidFileException, BadZipFile, KeyError, OSError):
            # Backup ablegen und neu anlegen
            backup = EXCEL_PATH.with_suffix(".bad.xlsx")
            try:
                EXCEL_PATH.rename(backup)
            except Exception:
                pass
            wb, ws = _new_wb_with_header(keys)
    else:
        wb, ws = _new_wb_with_header(keys)

    if wb is None or ws is None:
        # Fallback, sollte nicht passieren
        wb, ws = _new_wb_with_header(keys)

    ws.append([row[k] for k in keys])
    wb.save(EXCEL_PATH)


async def verify_recaptcha(token: str) -> bool:
    """Prüft reCAPTCHA v3, falls Secret gesetzt; sonst True."""
    if not RECAPTCHA_SECRET or not token:
        return True
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.post(
                "https://www.google.com/recaptcha/api/siteverify",
                data={"secret": RECAPTCHA_SECRET, "response": token},
            )
        data = r.json()
        return bool(data.get("success")) and float(data.get("score", 0.0)) >= 0.4
    except Exception:
        return False

# ---------------------------
# FastAPI App
# ---------------------------
app = FastAPI()

# Index explizit ausliefern (kein StaticFiles-Mount auf "/", damit POST /apply nicht 405 wird)
@app.get("/")
def serve_index():
    return FileResponse(str(BASE_DIR / "index.html"))


@app.post("/apply")
async def apply(request: Request):
    form = await request.form()

    # Pflichtfelder prüfen (müssen mit index.html übereinstimmen)
    required = ["name", "email", "phone", "location", "start_date", "visa"]
    missing = [k for k in required if not form.get(k)]
    resume = form.get("resume")  # UploadFile
    if not resume or not getattr(resume, "filename", ""):
        missing.append("resume")
    if missing:
        return JSONResponse({"message": f"Missing fields: {', '.join(missing)}"}, status_code=400)

    # reCAPTCHA v3 (optional)
    token = form.get("recaptcha_token", "")
    if not await verify_recaptcha(token):
        return JSONResponse({"message": "reCAPTCHA verification failed."}, status_code=400)

    # Datei lesen + Limits + Typen
    content = await resume.read()
    if len(content) > MAX_UPLOAD_MB * 1024 * 1024:
        return JSONResponse({"message": f"File too large (>{MAX_UPLOAD_MB}MB)."}, status_code=400)

    ext = Path(resume.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return JSONResponse({"message": "Unsupported file type. Please upload PDF/DOC/DOCX."}, status_code=400)

    # Sicher speichern
    safe_name = os.path.basename(resume.filename)
    path = UPLOAD_DIR / safe_name
    path.write_bytes(content)

    # Mehrfachauswahl (Expertise) + weitere Felder
    exp: List[str] = form.getlist("expertise[]") if hasattr(form, "getlist") else []
    links = (form.get("links") or "").strip()
    why   = (form.get("info") or "").strip()

    # Excel-Log immer schreiben
    append_to_excel({
        "timestamp": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "name": form.get("name"),
        "email": form.get("email"),
        "phone": form.get("phone"),
        "location": form.get("location"),
        "start_date": form.get("start_date"),
        "visa": form.get("visa"),
        "expertise": ", ".join(exp),
        "links": links,
        "why_kedorion": why,
        "resume_filename": safe_name,
    })

    # Immer Erfolg zurückgeben (keine E-Mail)
    return JSONResponse({"message": "Application received. Thank you!"})
